"""
OpenSight Web API - Robust CS2 Demo Analyzer

FastAPI application for CS2 demo analysis with professional-grade metrics.

Provides:
- Demo analysis with HLTV 2.0 Rating, KAST%, TTD, Crosshair Placement
- Simple synchronous processing for reliability
- Clean error handling with full logging

Endpoints:
- GET /         HTML interface with drop-zone
- GET /health   Health check
- POST /analyze Upload and analyze demo file
"""

import hashlib
import logging
import os
import pickle
import tempfile
import time
import traceback
from dataclasses import is_dataclass
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

__version__ = "0.4.0"


# =============================================================================
# Request/Response Models
# =============================================================================


class PlayerCompareRequest(BaseModel):
    """Request body for player comparison endpoint."""

    player_a: str
    player_b: str


# Security constants
MAX_FILE_SIZE_MB = 500
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024
ALLOWED_EXTENSIONS = (".dem", ".dem.gz")

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# =============================================================================
# Caching Configuration
# =============================================================================

CACHE_DIR = Path(tempfile.gettempdir()) / "opensight_cache"
CACHE_MAX_AGE_HOURS = 24
CACHE_CLEANUP_INTERVAL_SECONDS = 3600  # Run cleanup at most once per hour

# Track last cleanup time to avoid running too frequently
_last_cache_cleanup = 0.0


def get_file_hash(file_path: Path) -> str:
    """
    Calculate SHA256 hash of a file.

    Reads the file in chunks to handle large demo files efficiently.

    Args:
        file_path: Path to the file to hash

    Returns:
        Hex digest of the file's SHA256 hash
    """
    sha256 = hashlib.sha256()
    with open(file_path, "rb") as f:
        while chunk := f.read(8192):
            sha256.update(chunk)
    return sha256.hexdigest()


def get_cache_path(file_hash: str) -> Path:
    """Get the cache file path for a given file hash."""
    return CACHE_DIR / f"{file_hash}.pkl"


def load_from_cache(file_hash: str) -> dict | None:
    """
    Load cached analysis result if it exists and is valid.

    Args:
        file_hash: SHA256 hash of the demo file

    Returns:
        Cached result dict, or None if not found/invalid
    """
    cache_path = get_cache_path(file_hash)
    if not cache_path.exists():
        return None

    try:
        with open(cache_path, "rb") as f:
            result = pickle.load(f)
        logger.info(f"Cache hit for hash {file_hash[:16]}...")
        return result
    except (pickle.PickleError, EOFError, OSError) as e:
        logger.warning(f"Failed to load cache {cache_path}: {e}")
        # Remove corrupted cache file
        try:
            cache_path.unlink()
        except OSError:
            pass
        return None


def save_to_cache(file_hash: str, result: dict) -> None:
    """
    Save analysis result to cache.

    Args:
        file_hash: SHA256 hash of the demo file
        result: Analysis result dictionary to cache
    """
    try:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cache_path = get_cache_path(file_hash)
        with open(cache_path, "wb") as f:
            pickle.dump(result, f, protocol=pickle.HIGHEST_PROTOCOL)
        logger.info(f"Cached result for hash {file_hash[:16]}...")
    except (pickle.PickleError, OSError) as e:
        logger.warning(f"Failed to save cache: {e}")


def cleanup_old_cache() -> None:
    """
    Remove cache files older than CACHE_MAX_AGE_HOURS.

    Only runs if enough time has passed since last cleanup to avoid
    excessive filesystem operations.
    """
    global _last_cache_cleanup

    now = time.time()
    if now - _last_cache_cleanup < CACHE_CLEANUP_INTERVAL_SECONDS:
        return  # Too soon since last cleanup

    _last_cache_cleanup = now

    if not CACHE_DIR.exists():
        return

    max_age_seconds = CACHE_MAX_AGE_HOURS * 3600
    cutoff_time = now - max_age_seconds
    removed_count = 0

    try:
        for cache_file in CACHE_DIR.glob("*.pkl"):
            try:
                if cache_file.stat().st_mtime < cutoff_time:
                    cache_file.unlink()
                    removed_count += 1
            except OSError:
                pass

        if removed_count > 0:
            logger.info(f"Cache cleanup: removed {removed_count} old files")
    except OSError as e:
        logger.warning(f"Cache cleanup failed: {e}")


# =============================================================================
# FastAPI Application Setup
# =============================================================================

app = FastAPI(
    title="OpenSight API",
    description="CS2 demo analyzer - professional-grade metrics powered by awpy",
    version=__version__,
)

# Enable CORS for all origins (development mode)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files directory
# Path is relative to where uvicorn runs (from /app in Docker, or project root locally)
STATIC_DIR = Path(__file__).parent / "static"
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


@app.on_event("startup")
async def validate_imports():
    """Validate critical imports at startup to catch errors early."""
    errors = []
    try:
        from opensight.analytics import DemoAnalyzer
    except ImportError as e:
        errors.append(f"analytics.DemoAnalyzer: {e}")

    try:
        from opensight.parser import DemoParser
    except ImportError as e:
        errors.append(f"parser.DemoParser: {e}")

    if errors:
        logger.error(f"Critical import errors at startup: {errors}")
        # Don't raise - allow app to start but log the errors
    else:
        logger.info("✅ All critical imports validated successfully")


# =============================================================================
# Helper Functions
# =============================================================================


def dataclass_to_dict(obj: Any) -> Any:
    """Recursively convert dataclasses to dicts for JSON serialization."""
    if is_dataclass(obj) and not isinstance(obj, type):
        result = {}
        for field_name in obj.__dataclass_fields__:
            value = getattr(obj, field_name)
            result[field_name] = dataclass_to_dict(value)
        return result
    elif isinstance(obj, dict):
        return {k: dataclass_to_dict(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [dataclass_to_dict(item) for item in obj]
    elif isinstance(obj, (int, float, str, bool, type(None))):
        return obj
    else:
        # Try to convert to string for unknown types
        return str(obj)


def player_stats_to_dict(player: Any) -> dict:
    """Convert PlayerMatchStats to a JSON-serializable dict."""
    return {
        "steam_id": str(player.steam_id),
        "name": player.name,
        "team": player.team,
        "kills": player.kills,
        "deaths": player.deaths,
        "assists": player.assists,
        "headshots": player.headshots,
        "total_damage": player.total_damage,
        "rounds_played": player.rounds_played,
        "kd_ratio": player.kd_ratio,
        "kd_diff": player.kd_diff,
        "adr": player.adr,
        "headshot_percentage": player.headshot_percentage,
        "kast_percentage": player.kast_percentage,
        "hltv_rating": player.hltv_rating,
        "impact_rating": player.impact_rating,
        "survival_rate": player.survival_rate,
        "kills_per_round": player.kills_per_round,
        "deaths_per_round": player.deaths_per_round,
        # TTD (Time to Damage)
        "ttd_median_ms": round(player.ttd_median_ms, 1) if player.ttd_median_ms else None,
        "ttd_mean_ms": round(player.ttd_mean_ms, 1) if player.ttd_mean_ms else None,
        "ttd_samples": len(player.ttd_values),
        "prefire_count": player.prefire_count,
        # Crosshair Placement
        "cp_median_error_deg": round(player.cp_median_error_deg, 1)
        if player.cp_median_error_deg
        else None,
        "cp_mean_error_deg": round(player.cp_mean_error_deg, 1)
        if player.cp_mean_error_deg
        else None,
        "cp_samples": len(player.cp_values),
        # Opening duels
        "opening_duel_wins": player.opening_duels.wins,
        "opening_duel_losses": player.opening_duels.losses,
        "opening_duel_attempts": player.opening_duels.attempts,
        "opening_duel_win_rate": player.opening_duels.win_rate,
        # Trades
        "kills_traded": player.trades.kills_traded,
        "deaths_traded": player.trades.deaths_traded,
        # Clutches
        "clutch_situations": player.clutches.total_situations,
        "clutch_wins": player.clutches.total_wins,
        "clutch_win_rate": player.clutches.win_rate,
        # Multi-kills
        "rounds_with_2k": player.multi_kills.rounds_with_2k,
        "rounds_with_3k": player.multi_kills.rounds_with_3k,
        "rounds_with_4k": player.multi_kills.rounds_with_4k,
        "rounds_with_5k": player.multi_kills.rounds_with_5k,
        # Utility
        "flashbangs_thrown": player.utility.flashbangs_thrown,
        "enemies_flashed": player.utility.enemies_flashed,
        "flash_assists": player.utility.flash_assists,
        "he_thrown": player.utility.he_thrown,
        "he_damage": player.utility.he_damage,
        # Weapon breakdown
        "weapon_kills": player.weapon_kills,
    }


# =============================================================================
# HTML Template
# =============================================================================

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>CS2 Demo Analyzer</title>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
            background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
            min-height: 100vh;
            color: #e4e4e4;
            padding: 2rem;
        }
        .container { max-width: 1000px; margin: 0 auto; }
        h1 {
            text-align: center;
            font-size: 2.5rem;
            margin-bottom: 0.5rem;
            background: linear-gradient(90deg, #ffd700, #ff8c00);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }
        .subtitle {
            text-align: center;
            color: #888;
            margin-bottom: 2rem;
        }
        .drop-zone {
            border: 3px dashed #4a5568;
            border-radius: 1rem;
            padding: 4rem 2rem;
            text-align: center;
            background: rgba(255,255,255,0.02);
            transition: all 0.3s ease;
            cursor: pointer;
        }
        .drop-zone:hover, .drop-zone.dragover {
            border-color: #ffd700;
            background: rgba(255,215,0,0.05);
        }
        .drop-zone p { font-size: 1.2rem; margin-bottom: 1rem; }
        .drop-zone small { color: #666; }
        #file-input { display: none; }
        .status {
            margin-top: 1.5rem;
            padding: 1rem;
            border-radius: 0.5rem;
            display: none;
        }
        .status.loading { display: block; background: rgba(59, 130, 246, 0.2); color: #60a5fa; }
        .status.error { display: block; background: rgba(239, 68, 68, 0.2); color: #f87171; }
        .status.success { display: block; background: rgba(34, 197, 94, 0.2); color: #4ade80; }
        #results {
            margin-top: 2rem;
            display: none;
        }
        #results.visible { display: block; }
        .result-card {
            background: rgba(255,255,255,0.05);
            border-radius: 0.75rem;
            padding: 1.5rem;
            margin-bottom: 1rem;
        }
        .result-card h2 { color: #ffd700; margin-bottom: 1rem; font-size: 1.3rem; }
        .stat-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 1rem;
        }
        .stat-item { text-align: center; }
        .stat-value { font-size: 1.5rem; font-weight: bold; color: #fff; }
        .stat-label { font-size: 0.85rem; color: #888; }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 1rem;
        }
        th, td {
            padding: 0.75rem;
            text-align: left;
            border-bottom: 1px solid rgba(255,255,255,0.1);
        }
        th { color: #ffd700; font-weight: 600; }
        tr:hover { background: rgba(255,255,255,0.03); }
        .rating { font-weight: bold; }
        .rating.high { color: #4ade80; }
        .rating.medium { color: #fbbf24; }
        .rating.low { color: #f87171; }
        /* Tab styles */
        .tabs {
            display: flex;
            gap: 0.5rem;
            margin-bottom: 1rem;
            border-bottom: 2px solid rgba(255,255,255,0.1);
            padding-bottom: 0.5rem;
        }
        .tab-btn {
            background: rgba(255,255,255,0.05);
            border: none;
            color: #888;
            padding: 0.75rem 1.5rem;
            border-radius: 0.5rem 0.5rem 0 0;
            cursor: pointer;
            font-size: 0.95rem;
            transition: all 0.2s ease;
        }
        .tab-btn:hover { background: rgba(255,255,255,0.1); color: #fff; }
        .tab-btn.active { background: rgba(255,215,0,0.2); color: #ffd700; }
        .tab-content { display: none; }
        .tab-content.active { display: block; }
        /* Economy chart styles */
        .chart-container {
            position: relative;
            height: 350px;
            margin-top: 1rem;
            background: rgba(0,0,0,0.2);
            border-radius: 0.5rem;
            padding: 1rem;
        }
        .economy-legend {
            display: flex;
            justify-content: center;
            gap: 2rem;
            margin-top: 1rem;
        }
        .legend-item {
            display: flex;
            align-items: center;
            gap: 0.5rem;
            font-size: 0.9rem;
        }
        .legend-color {
            width: 20px;
            height: 4px;
            border-radius: 2px;
        }
        .legend-color.t-side { background: #f59e0b; }
        .legend-color.ct-side { background: #3b82f6; }
    </style>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
</head>
<body>
    <div class="container">
        <h1>CS2 Demo Analyzer</h1>
        <p class="subtitle">Powered by awpy - Professional-grade CS2 analytics</p>

        <div class="drop-zone" id="drop-zone">
            <p>Drop your .dem file here or click to select</p>
            <small>Supports .dem and .dem.gz files up to 500MB</small>
            <input type="file" id="file-input" accept=".dem,.dem.gz">
        </div>

        <div class="status" id="status"></div>

        <div id="results">
            <div class="result-card" id="match-info"></div>

            <div class="tabs" id="results-tabs">
                <button class="tab-btn active" data-tab="stats">Player Stats</button>
                <button class="tab-btn" data-tab="economy">Economy</button>
            </div>

            <div id="tab-stats" class="tab-content active">
                <div class="result-card" id="player-stats"></div>
            </div>

            <div id="tab-economy" class="tab-content">
                <div class="result-card">
                    <h2>Round-by-Round Economy</h2>
                    <p style="color: #888; margin-bottom: 1rem;">Equipment value per round - Identify eco rounds and force buys</p>
                    <div class="chart-container">
                        <canvas id="economyChart"></canvas>
                    </div>
                    <div class="economy-legend">
                        <div class="legend-item">
                            <div class="legend-color t-side"></div>
                            <span>Terrorists</span>
                        </div>
                        <div class="legend-item">
                            <div class="legend-color ct-side"></div>
                            <span>Counter-Terrorists</span>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <script>
        const dropZone = document.getElementById('drop-zone');
        const fileInput = document.getElementById('file-input');
        const status = document.getElementById('status');
        const results = document.getElementById('results');
        const matchInfo = document.getElementById('match-info');
        const playerStats = document.getElementById('player-stats');

        // Chart instance for cleanup
        let economyChart = null;

        function escapeHtml(text) {
            const div = document.createElement('div');
            div.textContent = text;
            return div.innerHTML;
        }

        // Tab switching
        document.querySelectorAll('.tab-btn').forEach(btn => {
            btn.addEventListener('click', () => {
                document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
                document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
                btn.classList.add('active');
                document.getElementById('tab-' + btn.dataset.tab).classList.add('active');
            });
        });

        dropZone.addEventListener('click', () => fileInput.click());
        dropZone.addEventListener('dragover', (e) => { e.preventDefault(); dropZone.classList.add('dragover'); });
        dropZone.addEventListener('dragleave', () => dropZone.classList.remove('dragover'));
        dropZone.addEventListener('drop', (e) => {
            e.preventDefault();
            dropZone.classList.remove('dragover');
            if (e.dataTransfer.files.length) handleFile(e.dataTransfer.files[0]);
        });
        fileInput.addEventListener('change', () => { if (fileInput.files.length) handleFile(fileInput.files[0]); });

        async function handleFile(file) {
            if (!file.name.match(/\\.dem(\\.gz)?$/i)) {
                showStatus('Please select a .dem or .dem.gz file', 'error');
                return;
            }
            showStatus('Analyzing demo... This may take a minute for large files.', 'loading');
            results.classList.remove('visible');

            const formData = new FormData();
            formData.append('file', file);

            try {
                const response = await fetch('/analyze', { method: 'POST', body: formData });
                const data = await response.json();

                if (!response.ok) {
                    showStatus(data.error || data.detail || 'Analysis failed', 'error');
                    return;
                }

                showStatus('Analysis complete!', 'success');
                displayResults(data);
            } catch (err) {
                showStatus('Network error: ' + err.message, 'error');
            }
        }

        function showStatus(message, type) {
            status.textContent = message;
            status.className = 'status ' + type;
        }

        function getRatingClass(rating) {
            if (rating >= 1.15) return 'high';
            if (rating >= 0.85) return 'medium';
            return 'low';
        }

        function displayResults(data) {
            matchInfo.innerHTML = `
                <h2>Match Information</h2>
                <div class="stat-grid">
                    <div class="stat-item">
                        <div class="stat-value">${escapeHtml(data.map_name)}</div>
                        <div class="stat-label">Map</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-value">${data.tick_rate}</div>
                        <div class="stat-label">Tick Rate</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-value">${data.total_rounds || '-'}</div>
                        <div class="stat-label">Rounds</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-value">${data.players ? data.players.length : 0}</div>
                        <div class="stat-label">Players</div>
                    </div>
                </div>
            `;

            if (data.players && data.players.length > 0) {
                let tableHtml = `
                    <h2>Player Statistics</h2>
                    <table>
                        <thead>
                            <tr>
                                <th>Player</th>
                                <th>Team</th>
                                <th>K/D/A</th>
                                <th>ADR</th>
                                <th>HS%</th>
                                <th>KAST%</th>
                                <th>Rating</th>
                            </tr>
                        </thead>
                        <tbody>
                `;

                data.players.sort((a, b) => b.hltv_rating - a.hltv_rating);

                for (const p of data.players) {
                    const ratingClass = getRatingClass(p.hltv_rating);
                    tableHtml += `
                        <tr>
                            <td>${escapeHtml(p.name)}</td>
                            <td>${escapeHtml(p.team)}</td>
                            <td>${p.kills}/${p.deaths}/${p.assists}</td>
                            <td>${p.adr.toFixed(1)}</td>
                            <td>${p.headshot_percentage.toFixed(1)}%</td>
                            <td>${p.kast_percentage.toFixed(1)}%</td>
                            <td class="rating ${ratingClass}">${p.hltv_rating.toFixed(2)}</td>
                        </tr>
                    `;
                }

                tableHtml += '</tbody></table>';
                playerStats.innerHTML = tableHtml;
            } else {
                playerStats.innerHTML = '<h2>Player Statistics</h2><p>No player data available.</p>';
            }

            results.classList.add('visible');

            // Create economy chart if data available
            if (data.economy_history && data.economy_history.length > 0) {
                createEconomyChart(data.economy_history);
            }
        }

        function createEconomyChart(economyData) {
            // Destroy existing chart if any
            if (economyChart) {
                economyChart.destroy();
            }

            const ctx = document.getElementById('economyChart').getContext('2d');

            const rounds = economyData.map(d => d.round);
            const tValues = economyData.map(d => d.team_t_val);
            const ctValues = economyData.map(d => d.team_ct_val);

            economyChart = new Chart(ctx, {
                type: 'line',
                data: {
                    labels: rounds,
                    datasets: [
                        {
                            label: 'Terrorists',
                            data: tValues,
                            borderColor: '#f59e0b',
                            backgroundColor: 'rgba(245, 158, 11, 0.1)',
                            borderWidth: 2,
                            fill: true,
                            tension: 0.3,
                            pointRadius: 4,
                            pointHoverRadius: 6
                        },
                        {
                            label: 'Counter-Terrorists',
                            data: ctValues,
                            borderColor: '#3b82f6',
                            backgroundColor: 'rgba(59, 130, 246, 0.1)',
                            borderWidth: 2,
                            fill: true,
                            tension: 0.3,
                            pointRadius: 4,
                            pointHoverRadius: 6
                        }
                    ]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    interaction: {
                        intersect: false,
                        mode: 'index'
                    },
                    plugins: {
                        legend: {
                            display: false  // Using custom legend
                        },
                        tooltip: {
                            backgroundColor: 'rgba(0, 0, 0, 0.8)',
                            titleColor: '#ffd700',
                            bodyColor: '#fff',
                            padding: 12,
                            callbacks: {
                                title: function(context) {
                                    return 'Round ' + context[0].label;
                                },
                                label: function(context) {
                                    const idx = context.dataIndex;
                                    const buyType = context.datasetIndex === 0
                                        ? economyData[idx].t_buy
                                        : economyData[idx].ct_buy;
                                    return context.dataset.label + ': $' +
                                        context.parsed.y.toLocaleString() +
                                        ' (' + buyType + ')';
                                }
                            }
                        }
                    },
                    scales: {
                        x: {
                            title: {
                                display: true,
                                text: 'Round',
                                color: '#888'
                            },
                            ticks: { color: '#888' },
                            grid: { color: 'rgba(255, 255, 255, 0.05)' }
                        },
                        y: {
                            title: {
                                display: true,
                                text: 'Equipment Value ($)',
                                color: '#888'
                            },
                            ticks: {
                                color: '#888',
                                callback: function(value) {
                                    return '$' + (value / 1000).toFixed(0) + 'k';
                                }
                            },
                            grid: { color: 'rgba(255, 255, 255, 0.05)' },
                            min: 0,
                            max: 50000
                        }
                    }
                }
            });
        }
    </script>
</body>
</html>
"""


# =============================================================================
# Endpoints
# =============================================================================


@app.get("/", response_class=HTMLResponse)
async def root():
    """Serve the main web interface with drop-zone."""
    # Prefer the feature-rich static index.html if available
    static_index = STATIC_DIR / "index.html"
    if static_index.exists():
        return HTMLResponse(content=static_index.read_text(), status_code=200)
    # Fallback to embedded template
    return HTMLResponse(content=HTML_TEMPLATE, status_code=200)


@app.get("/health")
async def health():
    """Health check endpoint."""
    return {"status": "ok"}


@app.post("/analyze")
async def analyze_demo(file: UploadFile = File(...)):
    """
    Analyze a CS2 demo file and return player statistics.

    Accepts .dem and .dem.gz files up to 500MB.
    Returns JSON with map_name, tick_rate, and players list.
    """
    tmp_path = None

    try:
        # Validate filename
        if not file.filename:
            raise HTTPException(status_code=400, detail="No filename provided")

        filename_lower = file.filename.lower()
        if not filename_lower.endswith(ALLOWED_EXTENSIONS):
            raise HTTPException(
                status_code=400,
                detail=f"Invalid file type. Allowed: .dem, .dem.gz. Got: {file.filename}",
            )

        # Read file content and check size
        content = await file.read()
        file_size_bytes = len(content)

        if file_size_bytes == 0:
            raise HTTPException(status_code=400, detail="Empty file uploaded")

        if file_size_bytes > MAX_FILE_SIZE_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"File too large: {file_size_bytes / (1024 * 1024):.1f}MB. Maximum: {MAX_FILE_SIZE_MB}MB",
            )

        logger.info(f"Received file: {file.filename} ({file_size_bytes / (1024 * 1024):.1f}MB)")

        # Write to temporary file
        suffix = ".dem.gz" if filename_lower.endswith(".dem.gz") else ".dem"
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(content)
            tmp_path = Path(tmp.name)

        logger.info(f"Wrote temp file: {tmp_path}")

        # Calculate file hash for caching
        file_hash = get_file_hash(tmp_path)
        logger.info(f"File hash: {file_hash[:16]}...")

        # Run periodic cache cleanup (non-blocking, runs at most once per hour)
        cleanup_old_cache()

        # Check for cached result
        cached_result = load_from_cache(file_hash)
        if cached_result is not None:
            logger.info("Returning cached result")
            return cached_result

        # Import and run analysis (cache miss)
        from opensight.analytics import (
            DemoAnalyzer,
            calculate_economy_history,
            compute_kill_positions,
            compute_utility_metrics,
        )
        from opensight.parser import DemoParser

        # Parse the demo
        logger.info("Starting demo parsing...")
        parser = DemoParser(tmp_path)
        match_data = parser.parse()
        logger.info(
            f"Parsed: {len(match_data.kills)} kills, {len(match_data.damages)} damages, {match_data.num_rounds} rounds"
        )

        # Run analytics
        logger.info("Starting analytics...")
        analyzer = DemoAnalyzer(match_data)
        analysis = analyzer.analyze()
        logger.info(f"Analysis complete: {len(analysis.players)} players")

        # Build response
        players_list = [player_stats_to_dict(player) for player in analysis.get_leaderboard()]

        response = {
            "map_name": analysis.map_name,
            "tick_rate": match_data.tick_rate,
            "total_rounds": analysis.total_rounds,
            "team1_score": analysis.team1_score,
            "team2_score": analysis.team2_score,
            "players": players_list,
        }

        # Kill Map data for radar visualization (detailed kill positions)
        response["kill_map"] = compute_kill_positions(match_data)

        # Grenade trajectory data for utility visualization (limit to 1000 positions)
        response["grenade_data"] = {
            "positions": analysis.grenade_positions[:1000],
            "team_stats": analysis.grenade_team_stats,
        }

        # Utility stats per player (Scope.gg style nade stats)
        utility_metrics = compute_utility_metrics(match_data)
        response["utility_stats"] = [metrics.to_dict() for metrics in utility_metrics.values()]

        # AI Coaching insights
        response["coaching"] = analysis.coaching_insights

        # Economy history for round-by-round visualization
        response["economy_history"] = calculate_economy_history(match_data)

        # Save to cache for future requests
        save_to_cache(file_hash, response)

        return response

    except HTTPException:
        # Re-raise HTTP exceptions as-is
        raise

    except Exception as e:
        # Log full traceback
        tb = traceback.format_exc()
        logger.error(f"Analysis failed: {e}\n{tb}")

        # Return error response
        return JSONResponse(status_code=500, content={"error": "Analysis failed", "detail": str(e)})

    finally:
        # Always clean up temporary file
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
                logger.info(f"Cleaned up temp file: {tmp_path}")
            except OSError as e:
                logger.warning(f"Failed to clean up temp file {tmp_path}: {e}")


# =============================================================================
# Additional Endpoints (for compatibility)
# =============================================================================


@app.get("/readiness")
async def readiness():
    """
    Readiness check for container orchestration.

    Checks disk space, temp directory, and core dependencies.
    """
    import shutil

    checks = {}
    all_ready = True

    # Check disk space
    try:
        temp_dir = tempfile.gettempdir()
        disk_usage = shutil.disk_usage(temp_dir)
        free_mb = disk_usage.free / (1024 * 1024)
        checks["disk_space"] = {
            "status": "ok" if free_mb >= 100 else "fail",
            "free_mb": round(free_mb),
        }
        if free_mb < 100:
            all_ready = False
    except Exception as e:
        checks["disk_space"] = {"status": "fail", "error": str(e)}
        all_ready = False

    # Check temp directory writable
    try:
        temp_dir = tempfile.gettempdir()
        test_file = os.path.join(temp_dir, f"opensight_test_{os.getpid()}.tmp")
        with open(test_file, "w") as f:
            f.write("test")
        os.remove(test_file)
        checks["temp_writable"] = {"status": "ok", "path": temp_dir}
    except Exception as e:
        checks["temp_writable"] = {"status": "fail", "error": str(e)}
        all_ready = False

    # Check dependencies
    dep_status = {}
    for dep in ["awpy", "pandas", "numpy"]:
        try:
            __import__(dep)
            dep_status[dep] = "ok"
        except ImportError as e:
            dep_status[dep] = f"fail: {e}"
            all_ready = False
    checks["dependencies"] = dep_status

    response = {"ready": all_ready, "version": __version__, "checks": checks}
    return JSONResponse(content=response, status_code=200 if all_ready else 503)


@app.post("/compare")
async def compare_players_endpoint(
    file: UploadFile = File(...),
    player_a: str = Query(..., description="Name of the first player to compare"),
    player_b: str = Query(..., description="Name of the second player to compare"),
):
    """
    Compare two players from a demo file using Scope.gg-style radar chart.

    This endpoint accepts a demo file and two player names, parses the demo,
    computes all metrics, and returns comparison data suitable for radar chart
    visualization.

    The comparison includes 5 axes:
    - ADR: Average Damage per Round
    - Opening Success %: Percentage of opening duels won
    - Clutch Win %: Percentage of clutch situations won
    - Trade Success %: Percentage of trade opportunities converted
    - Utility Usage: Grenades thrown per round (scaled)

    Returns normalized scores (0-100) for radar chart plus raw values for display.
    """
    # Validate file extension
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    filename_lower = file.filename.lower()
    if not filename_lower.endswith(ALLOWED_EXTENSIONS):
        raise HTTPException(
            status_code=400, detail=f"File must be a .dem or .dem.gz file. Got: {file.filename}"
        )

    # Verify analysis modules are available
    try:
        from opensight.analytics import DemoAnalyzer, compare_players
        from opensight.parser import DemoParser
    except ImportError as e:
        raise HTTPException(
            status_code=503, detail=f"Demo analysis not available. Missing: {str(e)}"
        )

    # Read and validate file
    content = await file.read()
    file_size_bytes = len(content)
    file_size_mb = file_size_bytes / (1024 * 1024)

    if file_size_bytes > MAX_FILE_SIZE_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File too large: {file_size_mb:.1f}MB. Maximum allowed: {MAX_FILE_SIZE_MB}MB",
        )

    if file_size_bytes == 0:
        raise HTTPException(status_code=400, detail="Empty file uploaded")

    tmp_path = None
    try:
        # Save file to temp location
        suffix = ".dem.gz" if filename_lower.endswith(".dem.gz") else ".dem"
        with NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(content)
            tmp_path = Path(tmp.name)

        # Parse the demo
        parser = DemoParser(tmp_path)
        data = parser.parse()

        # Run analytics
        analyzer = DemoAnalyzer(data)
        analysis = analyzer.analyze()

        # Perform player comparison
        try:
            comparison = compare_players(analysis, player_a, player_b)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

        # Return comparison data
        return JSONResponse(
            content={
                "status": "success",
                "demo_info": {
                    "map": analysis.map_name,
                    "rounds": analysis.total_rounds,
                    "score": f"{analysis.team1_score} - {analysis.team2_score}",
                },
                "comparison": comparison,
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Compare endpoint failed")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        # Clean up temp file
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


# NOTE: /compare/{job_id} endpoint disabled - requires job_store infrastructure
# which is not currently implemented. Use POST /compare with file upload instead.
# @app.post("/compare/{job_id}")
# async def compare_players_from_job(job_id: str, request: PlayerCompareRequest):
#     """Compare two players using cached job results."""
#     pass  # Requires job_store and JobStatus to be implemented


# Legacy code preserved for reference when job infrastructure is added:
def _compare_players_from_job_disabled(job_id: str, request: PlayerCompareRequest):
    """
    Compare two players using data from a completed analysis job.

    This is more efficient than re-analyzing the demo if you've already
    run /analyze - it uses the cached job results.

    Args:
        job_id: The job ID from a completed /analyze request
        request: JSON body with player_a and player_b names

    Returns normalized scores (0-100) for radar chart plus raw values for display.
    """
    # NOTE: Requires job_store and JobStatus to be implemented
    # job = job_store.get_job(job_id)
    # if not job:
    #     raise HTTPException(status_code=404, detail=f"Job not found: {job_id}")
    #
    # if job.status != JobStatus.COMPLETED:
    #     raise HTTPException(
    #         status_code=400,
    #         detail=f"Job not completed. Current status: {job.status.value}"
    #     )
    raise HTTPException(status_code=501, detail="Job-based comparison not implemented")


# =============================================================================
# Excel Export Endpoint
# =============================================================================


@app.post("/export/excel")
async def export_to_excel(file: UploadFile = File(...)):
    """
    Export CS2 demo analysis to a professionally formatted Excel file.

    Creates a multi-sheet Excel workbook with:
    - Overview: Player stats table (Name, K, D, ADR, Rating, etc.)
    - Kills: Detailed list of every kill event
    - Damages: Detailed list of damage events
    - Advanced: TTD, Clutch, Utility stats per player

    Returns a downloadable .xlsx file.
    """
    import io

    # Validate file extension
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    filename_lower = file.filename.lower()
    if not filename_lower.endswith(ALLOWED_EXTENSIONS):
        raise HTTPException(
            status_code=400, detail=f"File must be a .dem or .dem.gz file. Got: {file.filename}"
        )

    # Verify required modules
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        from opensight.analytics import DemoAnalyzer
        from opensight.parser import DemoParser
    except ImportError as e:
        raise HTTPException(
            status_code=503, detail=f"Excel export not available. Missing: {str(e)}"
        )

    # Read and validate file
    content = await file.read()
    file_size_bytes = len(content)
    file_size_mb = file_size_bytes / (1024 * 1024)

    if file_size_bytes > MAX_FILE_SIZE_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File too large: {file_size_mb:.1f}MB. Maximum allowed: {MAX_FILE_SIZE_MB}MB",
        )

    if file_size_bytes == 0:
        raise HTTPException(status_code=400, detail="Empty file uploaded")

    tmp_path = None
    try:
        # Save to temp file
        suffix = ".dem.gz" if filename_lower.endswith(".dem.gz") else ".dem"
        with NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(content)
            tmp_path = Path(tmp.name)

        # Parse the demo
        parser = DemoParser(tmp_path)
        data = parser.parse()

        # Run analytics
        analyzer = DemoAnalyzer(data)
        analysis = analyzer.analyze()

        # Create Excel workbook
        wb = Workbook()

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        ct_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        t_fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")

        def style_header_row(ws, num_cols):
            """Apply header styling to first row."""
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

        def add_autofilter(ws, num_cols):
            """Add autofilter to the worksheet."""
            if ws.max_row > 1:
                ws.auto_filter.ref = f"A1:{chr(64 + min(num_cols, 26))}{ws.max_row}"

        def adjust_column_widths(ws):
            """Auto-adjust column widths based on content."""
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = max(adjusted_width, 10)

        # =====================================================================
        # Sheet 1: Overview - Player Stats
        # =====================================================================
        ws_overview = wb.active
        ws_overview.title = "Overview"

        # Create player stats dataframe
        overview_data = []
        for player in analysis.get_leaderboard():
            overview_data.append(
                {
                    "Name": player.name,
                    "Team": player.team,
                    "K": player.kills,
                    "D": player.deaths,
                    "A": player.assists,
                    "K/D": round(player.kd_ratio, 2),
                    "+/-": player.kd_diff,
                    "ADR": round(player.adr, 1),
                    "HS%": round(player.headshot_percentage, 1),
                    "KAST%": round(player.kast_percentage, 1),
                    "Rating": round(player.hltv_rating, 2),
                    "Impact": round(player.impact_rating, 2),
                    "Aim Rating": round(player.aim_rating, 1) if player.aim_rating else 0,
                    "Utility Rating": round(player.utility_rating, 1)
                    if player.utility_rating
                    else 0,
                }
            )

        df_overview = pd.DataFrame(overview_data)

        # Write dataframe to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df_overview, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_overview.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                # Color-code by team
                if r_idx > 1 and c_idx == 2:  # Team column
                    if value == "CT":
                        for col in range(1, len(row) + 1):
                            ws_overview.cell(row=r_idx, column=col).fill = ct_fill
                    elif value == "T":
                        for col in range(1, len(row) + 1):
                            ws_overview.cell(row=r_idx, column=col).fill = t_fill

        style_header_row(ws_overview, len(df_overview.columns))
        add_autofilter(ws_overview, len(df_overview.columns))
        adjust_column_widths(ws_overview)

        # Add match info at the top (insert rows)
        ws_overview.insert_rows(1, 4)
        ws_overview.cell(row=1, column=1, value="Map:").font = Font(bold=True)
        ws_overview.cell(row=1, column=2, value=analysis.map_name)
        ws_overview.cell(row=2, column=1, value="Score:").font = Font(bold=True)
        ws_overview.cell(row=2, column=2, value=f"{analysis.team1_score} - {analysis.team2_score}")
        ws_overview.cell(row=3, column=1, value="Rounds:").font = Font(bold=True)
        ws_overview.cell(row=3, column=2, value=analysis.total_rounds)
        ws_overview.cell(row=3, column=3, value="Duration:").font = Font(bold=True)
        ws_overview.cell(row=3, column=4, value=f"{round(data.duration_seconds / 60, 1)} min")

        # =====================================================================
        # Sheet 2: Kills - Detailed Kill Events
        # =====================================================================
        ws_kills = wb.create_sheet("Kills")

        kills_data = []
        for kill in data.kills:
            kills_data.append(
                {
                    "Round": kill.round_num,
                    "Tick": kill.tick,
                    "Attacker": kill.attacker_name,
                    "Attacker Team": kill.attacker_side,
                    "Victim": kill.victim_name,
                    "Victim Team": kill.victim_side,
                    "Weapon": kill.weapon,
                    "Headshot": "Yes" if kill.headshot else "No",
                    "Wallbang": "Yes" if kill.wallbang else "No",
                    "Through Smoke": "Yes" if kill.through_smoke else "No",
                    "No Scope": "Yes" if kill.no_scope else "No",
                    "Blind Kill": "Yes" if kill.blind_kill else "No",
                    "Assister": kill.assister_name or "",
                }
            )

        if kills_data:
            df_kills = pd.DataFrame(kills_data)
            for r_idx, row in enumerate(dataframe_to_rows(df_kills, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_kills.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border

            style_header_row(ws_kills, len(df_kills.columns))
            add_autofilter(ws_kills, len(df_kills.columns))
            adjust_column_widths(ws_kills)
        else:
            ws_kills.cell(row=1, column=1, value="No kill data available")

        # =====================================================================
        # Sheet 3: Damages - Detailed Damage Events
        # =====================================================================
        ws_damages = wb.create_sheet("Damages")

        damages_data = []
        for dmg in data.damages[:5000]:  # Limit to 5000 rows to prevent huge files
            damages_data.append(
                {
                    "Round": dmg.round_num,
                    "Tick": dmg.tick,
                    "Attacker": dmg.attacker_name,
                    "Attacker Team": dmg.attacker_side,
                    "Victim": dmg.victim_name,
                    "Victim Team": dmg.victim_side,
                    "Weapon": dmg.weapon,
                    "Damage": dmg.damage,
                    "Damage Armor": dmg.damage_armor,
                    "Hitgroup": dmg.hitgroup,
                }
            )

        if damages_data:
            df_damages = pd.DataFrame(damages_data)
            for r_idx, row in enumerate(dataframe_to_rows(df_damages, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_damages.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border

            style_header_row(ws_damages, len(df_damages.columns))
            add_autofilter(ws_damages, len(df_damages.columns))
            adjust_column_widths(ws_damages)
        else:
            ws_damages.cell(row=1, column=1, value="No damage data available")

        # =====================================================================
        # Sheet 4: Advanced - TTD, Clutch, Utility Stats
        # =====================================================================
        ws_advanced = wb.create_sheet("Advanced")

        advanced_data = []
        for player in analysis.get_leaderboard():
            advanced_data.append(
                {
                    "Name": player.name,
                    "Team": player.team,
                    # TTD Stats
                    "TTD Median (ms)": round(player.ttd_median_ms, 1)
                    if player.ttd_median_ms
                    else None,
                    "TTD Mean (ms)": round(player.ttd_mean_ms, 1) if player.ttd_mean_ms else None,
                    "TTD Samples": len(player.ttd_values),
                    "Prefire Kills": player.prefire_count,
                    # Crosshair Placement
                    "CP Median (deg)": round(player.cp_median_error_deg, 1)
                    if player.cp_median_error_deg
                    else None,
                    "CP Mean (deg)": round(player.cp_mean_error_deg, 1)
                    if player.cp_mean_error_deg
                    else None,
                    "CP Samples": len(player.cp_values),
                    # Opening Duels
                    "Opening Wins": player.opening_duels.wins,
                    "Opening Losses": player.opening_duels.losses,
                    "Opening Win%": round(player.opening_duels.win_rate, 1),
                    # Trades
                    "Kills Traded": player.trades.kills_traded,
                    "Deaths Traded": player.trades.deaths_traded,
                    # Clutches
                    "Clutch Attempts": player.clutches.total_situations,
                    "Clutch Wins": player.clutches.total_wins,
                    "1v1 W/A": f"{player.clutches.wins_1v1}/{player.clutches.situations_1v1}",
                    "1v2 W/A": f"{player.clutches.wins_1v2}/{player.clutches.situations_1v2}",
                    "1v3+ W/A": f"{player.clutches.wins_1v3 + player.clutches.wins_1v4 + player.clutches.wins_1v5}/{player.clutches.situations_1v3 + player.clutches.situations_1v4 + player.clutches.situations_1v5}",
                    # Multi-kills
                    "2K Rounds": player.multi_kills.rounds_with_2k,
                    "3K Rounds": player.multi_kills.rounds_with_3k,
                    "4K Rounds": player.multi_kills.rounds_with_4k,
                    "5K (Ace)": player.multi_kills.rounds_with_5k,
                    # Utility
                    "Flashes Thrown": player.utility.flashbangs_thrown,
                    "Smokes Thrown": player.utility.smokes_thrown,
                    "Enemies Flashed": player.utility.enemies_flashed,
                    "Flash Eff": round(player.utility.enemies_flashed_per_flash, 2),
                    "HE Thrown": player.utility.he_thrown,
                    "HE Damage": player.utility.he_damage,
                    "Molotov Thrown": player.utility.molotovs_thrown,
                    "Molotov Damage": player.utility.molotov_damage,
                    "Flash Assists": player.utility.flash_assists,
                }
            )

        df_advanced = pd.DataFrame(advanced_data)
        for r_idx, row in enumerate(dataframe_to_rows(df_advanced, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_advanced.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                # Color-code by team
                if r_idx > 1 and c_idx == 2:
                    if value == "CT":
                        for col in range(1, len(row) + 1):
                            ws_advanced.cell(row=r_idx, column=col).fill = ct_fill
                    elif value == "T":
                        for col in range(1, len(row) + 1):
                            ws_advanced.cell(row=r_idx, column=col).fill = t_fill

        style_header_row(ws_advanced, len(df_advanced.columns))
        add_autofilter(ws_advanced, len(df_advanced.columns))
        adjust_column_widths(ws_advanced)

        # =====================================================================
        # Sheet 5: Rounds - Round-by-Round Data
        # =====================================================================
        ws_rounds = wb.create_sheet("Rounds")

        rounds_data = []
        for round_info in data.rounds:
            round_kills = [k for k in data.kills if k.round_num == round_info.round_num]
            rounds_data.append(
                {
                    "Round": round_info.round_num,
                    "Winner": round_info.winner,
                    "Win Reason": round_info.reason,
                    "CT Score": round_info.ct_score,
                    "T Score": round_info.t_score,
                    "Kills": len(round_kills),
                    "Round Type": round_info.round_type or "unknown",
                }
            )

        if rounds_data:
            df_rounds = pd.DataFrame(rounds_data)
            for r_idx, row in enumerate(dataframe_to_rows(df_rounds, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_rounds.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border
                    # Color-code winner column
                    if r_idx > 1 and c_idx == 2:
                        if value == "CT":
                            cell.fill = ct_fill
                        elif value == "T":
                            cell.fill = t_fill

            style_header_row(ws_rounds, len(df_rounds.columns))
            add_autofilter(ws_rounds, len(df_rounds.columns))
            adjust_column_widths(ws_rounds)

        # Save workbook to bytes buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Generate filename
        map_name = analysis.map_name.replace(" ", "_") if analysis.map_name else "unknown"
        output_filename = f"{map_name}_analysis.xlsx"

        logger.info(f"Excel export complete: {output_filename}")

        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{output_filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Excel export failed")
        raise HTTPException(status_code=500, detail=f"Excel export failed: {str(e)}")

    finally:
        # Clean up temp file
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


@app.get("/about")
async def about():
    """API documentation and metric descriptions."""
    return {
        "name": "OpenSight",
        "version": __version__,
        "description": "CS2 demo analyzer powered by awpy",
        "endpoints": {
            "GET /": "Web interface with file upload",
            "GET /health": "Health check",
            "GET /readiness": "Readiness check for containers",
            "POST /analyze": "Upload and analyze demo file",
            "GET /about": "This documentation",
        },
        "metrics": {
            "hltv_rating": "HLTV 2.0 Rating (1.0 = average)",
            "kast_percentage": "Rounds with Kill/Assist/Survived/Traded",
            "adr": "Average Damage per Round",
            "ttd_median_ms": "Time to Damage - reaction speed",
            "cp_median_error_deg": "Crosshair Placement - aim accuracy",
        },
    }


# =============================================================================
# Match History and Player Profile Endpoints (FREE - SQLite database)
# =============================================================================


@app.get("/history")
async def get_match_history(
    limit: int = Query(default=20, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    map_name: str | None = Query(default=None),
    steam_id: str | None = Query(default=None),
):
    """
    Get match history with optional filters.

    All data stored locally in SQLite (FREE, no cloud services).

    Args:
        limit: Maximum matches to return (1-100)
        offset: Pagination offset
        map_name: Filter by map name
        steam_id: Filter by player Steam ID
    """
    try:
        from opensight.database import get_db

        db = get_db()
        matches = db.get_match_history(
            limit=limit, offset=offset, map_name=map_name, steam_id=steam_id
        )
        stats = db.get_global_stats()

        return {
            "status": "success",
            "matches": matches,
            "total_matches": stats.get("total_matches", 0),
            "pagination": {"limit": limit, "offset": offset},
        }
    except Exception as e:
        logger.error(f"Failed to get match history: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/history/{match_id}")
async def get_match_details(match_id: int):
    """
    Get detailed information about a specific match.

    Includes all player statistics and round data.
    """
    try:
        from opensight.database import get_db

        db = get_db()
        details = db.get_match_details(match_id)

        if not details:
            raise HTTPException(status_code=404, detail="Match not found")

        return {"status": "success", **details}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get match details: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/profile/{steam_id}")
async def get_player_profile(steam_id: str):
    """
    Get a player's career profile and statistics.

    Aggregates performance across all analyzed matches.
    """
    try:
        from opensight.database import get_db

        db = get_db()
        profile = db.get_player_profile(steam_id)

        if not profile:
            raise HTTPException(status_code=404, detail="Player profile not found")

        return {"status": "success", "profile": profile}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get player profile: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/profile/{steam_id}/matches")
async def get_player_matches(steam_id: str, limit: int = Query(default=20, ge=1, le=50)):
    """Get a player's recent match history."""
    try:
        from opensight.database import get_db

        db = get_db()
        matches = db.get_player_match_history(steam_id, limit=limit)

        return {"status": "success", "matches": matches, "count": len(matches)}
    except Exception as e:
        logger.error(f"Failed to get player matches: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/profile/{steam_id}/insights")
async def get_player_insights(steam_id: str):
    """
    Get comprehensive insights about a player's performance.

    Includes:
    - Performance trends (improving/declining)
    - Career milestones achieved
    - Strengths and weaknesses
    - Recommended focus areas
    """
    try:
        from opensight.profiles import get_player_insights

        insights = get_player_insights(steam_id)

        if not insights:
            raise HTTPException(
                status_code=404, detail="Not enough data for insights (need 3+ matches)"
            )

        return {"status": "success", "insights": insights}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get player insights: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/leaderboard")
async def get_leaderboard(
    metric: str = Query(default="avg_rating"),
    min_matches: int = Query(default=5, ge=1),
    limit: int = Query(default=20, ge=1, le=100),
):
    """
    Get player leaderboard by metric.

    Available metrics:
    - avg_rating: Average HLTV rating
    - avg_adr: Average damage per round
    - avg_kast: Average KAST percentage
    - total_kills: Total career kills
    - total_matches: Total matches played
    """
    try:
        from opensight.database import get_db

        valid_metrics = ["avg_rating", "avg_adr", "avg_kast", "total_kills", "total_matches"]
        if metric not in valid_metrics:
            raise HTTPException(
                status_code=400, detail=f"Invalid metric. Choose from: {valid_metrics}"
            )

        db = get_db()
        leaderboard = db.get_leaderboard(metric=metric, min_matches=min_matches, limit=limit)

        return {
            "status": "success",
            "metric": metric,
            "min_matches": min_matches,
            "leaderboard": leaderboard,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get leaderboard: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/players/search")
async def search_players(q: str = Query(..., min_length=2), limit: int = Query(default=10, ge=1, le=50)):
    """Search for players by name."""
    try:
        from opensight.database import get_db

        db = get_db()
        results = db.search_players(q, limit=limit)

        return {"status": "success", "query": q, "results": results}
    except Exception as e:
        logger.error(f"Failed to search players: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/stats")
async def get_global_stats():
    """
    Get global statistics across all analyzed matches.

    Returns total matches, players, rounds, and map distribution.
    """
    try:
        from opensight.database import get_db

        db = get_db()
        stats = db.get_global_stats()

        return {"status": "success", "stats": stats}
    except Exception as e:
        logger.error(f"Failed to get global stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# Team Performance Analysis Endpoints (FREE - local processing)
# =============================================================================


@app.post("/team-analysis")
async def analyze_team_performance(file: UploadFile = File(...)):
    """
    Analyze team performance metrics from an uploaded demo file.

    Returns comprehensive metrics for both CT and T teams including:
    - KDA ratios
    - Kill distances
    - Round win rates
    - Trade success rates
    - Player breakdowns
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    if not (file.filename.endswith(".dem") or file.filename.endswith(".dem.gz")):
        raise HTTPException(status_code=400, detail="Only .dem and .dem.gz files accepted")

    temp_path = None
    try:
        # Save uploaded file
        import tempfile

        with tempfile.NamedTemporaryFile(delete=False, suffix=".dem") as tmp:
            content = await file.read()
            if len(content) == 0:
                raise HTTPException(status_code=400, detail="Empty file uploaded")
            tmp.write(content)
            temp_path = tmp.name

        # Parse demo
        from opensight.parser import DemoParser

        parser = DemoParser()
        match_data = parser.parse(temp_path)

        # Calculate team metrics
        from opensight.team_performance_metrics import calculate_team_metrics

        analysis = calculate_team_metrics(match_data)

        return {
            "status": "success",
            "filename": file.filename,
            "analysis": analysis.to_dict(),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Team analysis failed: {e}")
        raise HTTPException(status_code=500, detail=f"Analysis failed: {str(e)}")
    finally:
        if temp_path:
            import os

            try:
                os.unlink(temp_path)
            except OSError:
                pass


@app.post("/team-comparison")
async def compare_teams_endpoint(file: UploadFile = File(...)):
    """
    Compare CT and T team performance from a demo file.

    Returns detailed comparison including:
    - Metric-by-metric comparison
    - Team strengths and weaknesses
    - Improvement recommendations
    - Visual chart data
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    if not (file.filename.endswith(".dem") or file.filename.endswith(".dem.gz")):
        raise HTTPException(status_code=400, detail="Only .dem and .dem.gz files accepted")

    temp_path = None
    try:
        import tempfile

        with tempfile.NamedTemporaryFile(delete=False, suffix=".dem") as tmp:
            content = await file.read()
            if len(content) == 0:
                raise HTTPException(status_code=400, detail="Empty file uploaded")
            tmp.write(content)
            temp_path = tmp.name

        # Parse and analyze
        from opensight.parser import DemoParser
        from opensight.team_comparison import compare_teams_from_match, generate_comparison_charts

        parser = DemoParser()
        match_data = parser.parse(temp_path)
        comparison = compare_teams_from_match(match_data)
        charts = generate_comparison_charts(comparison)

        return {
            "status": "success",
            "filename": file.filename,
            "comparison": comparison.to_dict(),
            "charts": charts,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Team comparison failed: {e}")
        raise HTTPException(status_code=500, detail=f"Comparison failed: {str(e)}")
    finally:
        if temp_path:
            import os

            try:
                os.unlink(temp_path)
            except OSError:
                pass


@app.post("/team-comparison/html")
async def get_team_comparison_html(file: UploadFile = File(...)):
    """
    Generate an HTML report comparing team performance.

    Returns a complete HTML page with charts and analysis.
    Can be saved and opened in any browser.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    if not (file.filename.endswith(".dem") or file.filename.endswith(".dem.gz")):
        raise HTTPException(status_code=400, detail="Only .dem and .dem.gz files accepted")

    temp_path = None
    try:
        import tempfile

        with tempfile.NamedTemporaryFile(delete=False, suffix=".dem") as tmp:
            content = await file.read()
            if len(content) == 0:
                raise HTTPException(status_code=400, detail="Empty file uploaded")
            tmp.write(content)
            temp_path = tmp.name

        # Parse and generate report
        from opensight.parser import DemoParser
        from opensight.team_comparison import compare_teams_from_match, generate_comparison_html

        parser = DemoParser()
        match_data = parser.parse(temp_path)
        comparison = compare_teams_from_match(match_data)
        html_report = generate_comparison_html(comparison)

        from fastapi.responses import HTMLResponse

        return HTMLResponse(content=html_report, media_type="text/html")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"HTML report generation failed: {e}")
        raise HTTPException(status_code=500, detail=f"Report generation failed: {str(e)}")
    finally:
        if temp_path:
            import os

            try:
                os.unlink(temp_path)
            except OSError:
                pass


# =============================================================================
# Development Server
# =============================================================================

if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "opensight.api:app",
        host="0.0.0.0",
        port=7860,
        reload=True,
        log_level="info",
    )
